from fastapi import APIRouter, HTTPException, Depends, UploadFile, File
from fastapi.responses import StreamingResponse, FileResponse
from pydantic import BaseModel
from typing import List, Optional
import datetime
import io
import os
import shutil
import mimetypes
from database import get_db_connection
import sqlite3
from auth import get_current_admin
from utils.email_helper import send_factura_email

RECIBOS_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..", "data", "recibos_gastos"))
ALLOWED_EXTENSIONS = {".jpg", ".jpeg", ".png", ".pdf"}

router = APIRouter(
    prefix="/api",
    tags=["facturas"]
)

# Pydantic schemas for request validation
class ItemFacturaCreate(BaseModel):
    producto_id: Optional[int] = None
    nombre_producto: str
    cantidad: int
    precio_unitario: float
    total: float

class FacturaCreate(BaseModel):
    numero_factura: Optional[str] = None
    cliente_id: Optional[int] = None
    cliente_nombre_manual: Optional[str] = None  # usado cuando no hay cliente_id (órdenes directas/cotización)
    fecha_emision: str
    fecha_vencimiento: Optional[str] = None
    fecha_pago: Optional[str] = None
    metodo_pago: Optional[str] = None
    numero_cheque: Optional[str] = None
    subtotal: float
    ivu_estatal: float
    ivu_municipal: float
    total: float
    monto_pagado: Optional[float] = None
    notas: Optional[str] = None
    estado: str = "Pendiente"
    items: List[ItemFacturaCreate]
    orden_produccion_id: Optional[int] = None   # vínculo formal con ordenes_produccion
    orden_ids: Optional[List[int]] = None       # todas las órdenes del pedido agrupado
    codigo_orden: Optional[str] = None          # copia desnormalizada del código para mostrar rápido
    cotizacion_id: Optional[int] = None         # cotización origen (si se creó desde cotización)

class EstadoUpdate(BaseModel):
    estado: str
    fecha_pago: Optional[str] = None
    metodo_pago: Optional[str] = None
    numero_cheque: Optional[str] = None
    monto_pagado: Optional[float] = None  # Monto abonado (puede ser parcial)

@router.get("/facturas")
def listar_facturas(current_user: dict = Depends(get_current_admin)):
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT f.*,
                   COALESCE(c.nombre, f.cliente_nombre_manual, 'Cliente Directo') as cliente_nombre,
                   COALESCE(c.email, '') as cliente_email,
                   f.orden_produccion_id,
                   f.codigo_orden
            FROM facturas f
            LEFT JOIN clientes c ON f.cliente_id = c.id
            ORDER BY f.fecha_emision DESC, f.id DESC
        """)
        facturas = [dict(row) for row in cursor.fetchall()]
        conn.close()
        return {"status": "success", "data": facturas}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al listar facturas: {str(e)}")

@router.get("/facturas/nuevas")
def obtener_facturas_nuevas(current_user: dict = Depends(get_current_admin)):
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT f.id, f.numero_factura, f.total,
                   COALESCE(c.nombre, f.cliente_nombre_manual, 'Cliente Directo') as cliente_nombre
            FROM facturas f
            LEFT JOIN clientes c ON f.cliente_id = c.id
            WHERE f.notificado = 0
        """)
        rows = cursor.fetchall()
        nuevas = [dict(row) for row in rows]
        conn.close()
        return {"status": "success", "data": nuevas}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/facturas/{factura_id}")
def obtener_factura(factura_id: int, current_user: dict = Depends(get_current_admin)):
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Obtener factura
        cursor.execute("""
            SELECT f.*,
                   COALESCE(c.nombre, f.cliente_nombre_manual, 'Cliente Directo') as cliente_nombre,
                   COALESCE(c.contacto, '') as cliente_contacto,
                   COALESCE(c.email, '') as cliente_email,
                   COALESCE(c.telefono, '') as cliente_telefono,
                   COALESCE(c.notas, '') as cliente_notas
            FROM facturas f
            LEFT JOIN clientes c ON f.cliente_id = c.id
            WHERE f.id = ?
        """, (factura_id,))
        factura_row = cursor.fetchone()
        
        if not factura_row:
            conn.close()
            raise HTTPException(status_code=404, detail="Factura no encontrada")
            
        factura = dict(factura_row)
        
        # Obtener partidas con foto, dimensiones y material principal
        cursor.execute("""
            SELECT i.*,
                   (SELECT pi.ruta_imagen FROM producto_imagenes pi
                    WHERE pi.producto_id = i.producto_id
                    ORDER BY pi.es_principal DESC, pi.orden ASC, pi.id ASC LIMIT 1) as foto_galeria,
                   (SELECT f.nombre_archivo FROM fotos_asociadas f
                    WHERE f.producto_id = i.producto_id AND f.tipo_foto = 'referencia'
                    ORDER BY f.id DESC LIMIT 1) as foto_nombre,
                   p.ancho, p.alto, p.tipo_producto,
                   (SELECT m.nombre FROM componentes_producto cp
                    JOIN materiales m ON cp.material_id = m.id
                    WHERE cp.producto_id = i.producto_id
                    ORDER BY cp.cantidad_usada DESC LIMIT 1) as material_principal
            FROM items_factura i
            LEFT JOIN productos p ON i.producto_id = p.id
            WHERE i.factura_id = ?
        """, (factura_id,))
        items = []
        for row in cursor.fetchall():
            d = dict(row)
            if d.get('foto_galeria'):
                d['foto_ruta'] = d['foto_galeria']
            elif d.get('foto_nombre'):
                d['foto_ruta'] = f"/fotos_import/{d['foto_nombre']}"
            else:
                d['foto_ruta'] = None
            items.append(d)
        factura["items"] = items
        
        conn.close()
        return {"status": "success", "data": factura}
    except HTTPException as he:
        raise he
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al obtener detalle de factura: {str(e)}")

@router.post("/facturas")
def crear_factura(factura: FacturaCreate, current_user: dict = Depends(get_current_admin)):
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Generar número de factura si no se provee
        num_factura = factura.numero_factura
        if not num_factura:
            anio_actual = datetime.datetime.now().strftime("%Y")
            cursor.execute("""
                SELECT numero_factura FROM facturas 
                WHERE numero_factura LIKE ? 
                ORDER BY numero_factura DESC LIMIT 1
            """, (f"EV-{anio_actual}-%",))
            last_row = cursor.fetchone()
            
            if last_row:
                last_num = last_row["numero_factura"]
                # EV-YYYY-XXXX
                try:
                    seq = int(last_num.split("-")[-1])
                    next_seq = seq + 1
                except ValueError:
                    next_seq = 1
            else:
                next_seq = 1
            num_factura = f"EV-{anio_actual}-{next_seq:04d}"
            
        # Verificar que el cliente existe (solo si se provee cliente_id)
        if factura.cliente_id is not None:
            cursor.execute("SELECT id FROM clientes WHERE id = ?", (factura.cliente_id,))
            if not cursor.fetchone():
                conn.close()
                raise HTTPException(status_code=404, detail="Cliente no encontrado")

        # Guard anti-duplicado: si alguna orden del pedido ya fue facturada, rechazar.
        ids_a_verificar = list(factura.orden_ids or [])
        if factura.orden_produccion_id and factura.orden_produccion_id not in ids_a_verificar:
            ids_a_verificar.append(factura.orden_produccion_id)

        if ids_a_verificar:
            marcadores = ",".join("?" for _ in ids_a_verificar)
            cursor.execute(f"""
                SELECT o.codigo_orden, f.numero_factura
                FROM ordenes_produccion o
                LEFT JOIN facturas f ON f.id = o.factura_id
                WHERE o.id IN ({marcadores}) AND o.factura_id IS NOT NULL
                LIMIT 1
            """, ids_a_verificar)
            ya_facturada = cursor.fetchone()
            if ya_facturada:
                conn.close()
                raise HTTPException(
                    status_code=409,
                    detail=(
                        f"La orden '{ya_facturada['codigo_orden']}' ya tiene la factura "
                        f"{ya_facturada['numero_factura'] or 'asociada'}. No se creó una factura duplicada."
                    )
                )

        # Iniciar transacción
        try:
            cursor.execute("""
                INSERT INTO facturas (
                    numero_factura, cliente_id, cliente_nombre_manual,
                    fecha_emision, fecha_vencimiento,
                    fecha_pago, metodo_pago, numero_cheque, subtotal,
                    ivu_estatal, ivu_municipal, total, monto_pagado, notas, estado,
                    orden_produccion_id, codigo_orden, cotizacion_id
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                num_factura, factura.cliente_id, factura.cliente_nombre_manual,
                factura.fecha_emision, factura.fecha_vencimiento,
                factura.fecha_pago, factura.metodo_pago, factura.numero_cheque, factura.subtotal,
                factura.ivu_estatal, factura.ivu_municipal, factura.total, factura.monto_pagado, factura.notas, factura.estado,
                factura.orden_produccion_id, factura.codigo_orden, factura.cotizacion_id
            ))
            
            factura_id = cursor.lastrowid
            
            # Insertar partidas
            for item in factura.items:
                prod_id = item.producto_id
                if not prod_id:
                    cursor.execute("SELECT id FROM productos WHERE nombre = ?", (item.nombre_producto,))
                    prod_row = cursor.fetchone()
                    if prod_row:
                        prod_id = prod_row["id"]
                cursor.execute("""
                    INSERT INTO items_factura (
                        factura_id, producto_id, nombre_producto, cantidad, precio_unitario, total
                    ) VALUES (?, ?, ?, ?, ?, ?)
                """, (
                    factura_id, prod_id, item.nombre_producto, item.cantidad, item.precio_unitario, item.total
                ))
                
            # Vincular TODAS las órdenes del pedido con esta factura (misma transacción).
            # Si solo se vinculara la primera, las demás quedarían sin factura_id y el
            # Kanban volvería a ofrecer "Facturar", creando facturas duplicadas.
            orden_ids = list(factura.orden_ids or [])
            if factura.orden_produccion_id and factura.orden_produccion_id not in orden_ids:
                orden_ids.append(factura.orden_produccion_id)

            for oid in orden_ids:
                cursor.execute(
                    "UPDATE ordenes_produccion SET factura_id = ? WHERE id = ?",
                    (factura_id, oid)
                )

            conn.commit()

            conn.close()
            return {
                "status": "success",
                "message": "Factura creada con éxito",
                "id": factura_id,
                "numero_factura": num_factura,
                "orden_produccion_id": factura.orden_produccion_id,
                "codigo_orden": factura.codigo_orden,
            }
        except sqlite3.IntegrityError as ie:
            conn.rollback()
            conn.close()
            raise HTTPException(status_code=400, detail=f"El número de factura '{num_factura}' ya existe o hay un error de integridad.")
        except Exception as e:
            conn.rollback()
            conn.close()
            raise HTTPException(status_code=500, detail=f"Error transaccional de base de datos: {str(e)}")
            
    except HTTPException as he:
        raise he
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al registrar factura: {str(e)}")

@router.put("/facturas/{factura_id}/estado")
def actualizar_estado_factura(factura_id: int, data: EstadoUpdate, current_user: dict = Depends(get_current_admin)):
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Verificar existencia
        cursor.execute("SELECT id FROM facturas WHERE id = ?", (factura_id,))
        if not cursor.fetchone():
            conn.close()
            raise HTTPException(status_code=404, detail="Factura no encontrada")
            
        # Si el estado es 'Pagada', nos aseguramos de registrar fecha_pago si no se provee
        fecha_pago_val = data.fecha_pago
        if data.estado == "Pagada" and not fecha_pago_val:
            fecha_pago_val = datetime.datetime.now().strftime("%Y-%m-%d")
            
        cursor.execute("""
            UPDATE facturas
            SET estado = ?,
                fecha_pago = ?,
                metodo_pago = ?,
                numero_cheque = ?,
                monto_pagado = ?
            WHERE id = ?
        """, (data.estado, fecha_pago_val, data.metodo_pago, data.numero_cheque, data.monto_pagado, factura_id))
        
        conn.commit()
        conn.close()
        return {"status": "success", "message": "Estado de factura actualizado"}
    except HTTPException as he:
        raise he
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al actualizar estado: {str(e)}")

@router.delete("/facturas/{factura_id}")
def eliminar_factura(factura_id: int, current_user: dict = Depends(get_current_admin)):
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Verificar existencia
        cursor.execute("SELECT id FROM facturas WHERE id = ?", (factura_id,))
        if not cursor.fetchone():
            conn.close()
            raise HTTPException(status_code=404, detail="Factura no encontrada")
            
        # Borrar (on delete cascade limpiará items_factura)
        cursor.execute("DELETE FROM facturas WHERE id = ?", (factura_id,))
        conn.commit()
        conn.close()
        return {"status": "success", "message": "Factura eliminada con éxito"}
    except HTTPException as he:
        raise he
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al eliminar factura: {str(e)}")

@router.get("/contabilidad/reporte")
def reporte_contabilidad(current_user: dict = Depends(get_current_admin)):
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Obtener resumen por meses de fecha_emision de facturas no anuladas (Ciclo: del 20 de un mes al 19 del próximo mes)
        cursor.execute("""
            SELECT 
                CASE 
                    WHEN CAST(strftime('%d', fecha_emision) AS INTEGER) >= 20 
                    THEN strftime('%Y-%m', date(fecha_emision, '+1 month'))
                    ELSE strftime('%Y-%m', fecha_emision)
                END as mes,
                COUNT(id) as total_facturas,
                SUM(CASE WHEN estado != 'Anulada' THEN subtotal ELSE 0.0 END) as ingresos_subtotal,
                SUM(CASE WHEN estado != 'Anulada' THEN ivu_estatal ELSE 0.0 END) as ivu_estatal,
                SUM(CASE WHEN estado != 'Anulada' THEN ivu_municipal ELSE 0.0 END) as ivu_municipal,
                SUM(CASE WHEN estado != 'Anulada' THEN total ELSE 0.0 END) as total_facturado,
                SUM(CASE WHEN estado = 'Pagada' THEN total ELSE 0.0 END) as total_recaudado,
                SUM(CASE WHEN estado = 'Pendiente' THEN total ELSE 0.0 END) as total_pendiente,
                SUM(CASE WHEN estado = 'Anulada' THEN total ELSE 0.0 END) as total_anulado
            FROM facturas
            GROUP BY mes
            ORDER BY mes DESC
        """)
        reporte_meses = [dict(row) for row in cursor.fetchall()]
        
        # Totales acumulados históricos de facturas no anuladas
        cursor.execute("""
            SELECT 
                COUNT(id) as total_facturas,
                SUM(CASE WHEN estado != 'Anulada' THEN subtotal ELSE 0.0 END) as ingresos_subtotal,
                SUM(CASE WHEN estado != 'Anulada' THEN ivu_estatal ELSE 0.0 END) as ivu_estatal,
                SUM(CASE WHEN estado != 'Anulada' THEN ivu_municipal ELSE 0.0 END) as ivu_municipal,
                SUM(CASE WHEN estado != 'Anulada' THEN total ELSE 0.0 END) as total_facturado,
                SUM(CASE WHEN estado = 'Pagada' THEN total ELSE 0.0 END) as total_recaudado,
                SUM(CASE WHEN estado = 'Pendiente' THEN total ELSE 0.0 END) as total_pendiente
            FROM facturas
        """)
        row = cursor.fetchone()
        totales = dict(row) if row else {
            "total_facturas": 0,
            "ingresos_subtotal": 0.0,
            "ivu_estatal": 0.0,
            "ivu_municipal": 0.0,
            "total_facturado": 0.0,
            "total_recaudado": 0.0,
            "total_pendiente": 0.0
        }
        
        # Rellenar valores None con 0.0 si es que está vacía
        for key in totales:
            if totales[key] is None:
                totales[key] = 0.0
        
        conn.close()
        return {
            "status": "success",
            "totales": totales,
            "reporte_mensual": reporte_meses
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al generar reporte contable: {str(e)}")

class MarcarLeidasSchema(BaseModel):
    ids: List[int]

@router.post("/facturas/marcar-leidas")
def marcar_facturas_leidas(payload: MarcarLeidasSchema, current_user: dict = Depends(get_current_admin)):
    try:
        if not payload.ids:
            return {"status": "success", "message": "No hay facturas que marcar."}
        conn = get_db_connection()
        cursor = conn.cursor()
        placeholders = ",".join("?" for _ in payload.ids)
        cursor.execute(f"""
            UPDATE facturas 
            SET notificado = 1 
            WHERE id IN ({placeholders})
        """, payload.ids)
        conn.commit()
        conn.close()
        return {"status": "success", "message": f"{cursor.rowcount} facturas marcadas como leídas."}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# --- MODELOS GASTOS ---
class GastoCreate(BaseModel):
    concepto: str
    categoria: str
    monto: float
    fecha: str
    metodo_pago: Optional[str] = None
    notas: Optional[str] = None

class GastoUpdate(BaseModel):
    concepto: Optional[str] = None
    categoria: Optional[str] = None
    monto: Optional[float] = None
    fecha: Optional[str] = None
    metodo_pago: Optional[str] = None
    notas: Optional[str] = None
    proveedor: Optional[str] = None   # campo extra para Fase 2B (IA)

# --- ENDPOINTS GASTOS ---
@router.get("/contabilidad/ivu-periodo")
def ivu_por_periodo(
    fecha_inicio: str,
    fecha_fin: str,
    current_user: dict = Depends(get_current_admin)
):
    """Calcula IVU devengado y resumen de ventas para un rango de fechas personalizado."""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT
                COUNT(id)                                                              AS total_facturas,
                SUM(CASE WHEN estado != 'Anulada' THEN subtotal      ELSE 0.0 END)   AS ventas_sujetas_ivu,
                SUM(CASE WHEN estado != 'Anulada' THEN ivu_estatal   ELSE 0.0 END)   AS ivu_estatal,
                SUM(CASE WHEN estado != 'Anulada' THEN ivu_municipal  ELSE 0.0 END)  AS ivu_municipal,
                SUM(CASE WHEN estado != 'Anulada' THEN total          ELSE 0.0 END)  AS total_facturado,
                SUM(CASE WHEN estado = 'Pagada'   THEN total          ELSE 0.0 END)  AS total_recaudado,
                SUM(CASE WHEN estado = 'Pendiente' THEN total         ELSE 0.0 END)  AS total_pendiente
            FROM facturas
            WHERE fecha_emision >= ? AND fecha_emision <= ?
        """, (fecha_inicio, fecha_fin))
        row = cursor.fetchone()
        conn.close()
        d = dict(row) if row else {}
        for key in d:
            if d[key] is None:
                d[key] = 0.0
        ivu_estatal   = d.get("ivu_estatal", 0.0)
        ivu_municipal = d.get("ivu_municipal", 0.0)
        return {
            "status": "success",
            "periodo": {"desde": fecha_inicio, "hasta": fecha_fin},
            "total_facturas":    d.get("total_facturas", 0),
            "ventas_sujetas_ivu": round(d.get("ventas_sujetas_ivu", 0.0), 2),
            "ivu_estatal":        round(ivu_estatal, 2),
            "ivu_municipal":      round(ivu_municipal, 2),
            "ivu_total":          round(ivu_estatal + ivu_municipal, 2),
            "total_facturado":    round(d.get("total_facturado", 0.0), 2),
            "total_recaudado":    round(d.get("total_recaudado", 0.0), 2),
            "total_pendiente":    round(d.get("total_pendiente", 0.0), 2),
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al calcular IVU por periodo: {str(e)}")

@router.get("/contabilidad/ivu-periodo/exportar")
def exportar_ivu_excel(
    fecha_inicio: str,
    fecha_fin: str,
    current_user: dict = Depends(get_current_admin)
):
    """Genera un archivo .xlsx con el detalle de facturas y resumen IVU para el período indicado."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        conn = get_db_connection()
        cursor = conn.cursor()

        # Detalle de facturas en el período, con nombre del cliente
        cursor.execute("""
            SELECT
                f.numero_factura,
                f.fecha_emision,
                COALESCE(c.nombre, 'Cliente directo') AS cliente,
                f.subtotal,
                f.ivu_estatal,
                f.ivu_municipal,
                (f.ivu_estatal + f.ivu_municipal)   AS ivu_total,
                f.total,
                f.estado
            FROM facturas f
            LEFT JOIN clientes c ON f.cliente_id = c.id
            WHERE f.fecha_emision >= ? AND f.fecha_emision <= ?
            ORDER BY f.fecha_emision ASC, f.id ASC
        """, (fecha_inicio, fecha_fin))
        rows = cursor.fetchall()

        # Totales
        cursor.execute("""
            SELECT
                SUM(CASE WHEN estado != 'Anulada' THEN subtotal      ELSE 0 END) AS ventas_sujetas,
                SUM(CASE WHEN estado != 'Anulada' THEN ivu_estatal   ELSE 0 END) AS ivu_estatal,
                SUM(CASE WHEN estado != 'Anulada' THEN ivu_municipal ELSE 0 END) AS ivu_municipal,
                SUM(CASE WHEN estado != 'Anulada' THEN total         ELSE 0 END) AS total_facturado,
                SUM(CASE WHEN estado = 'Pagada'   THEN total         ELSE 0 END) AS total_recaudado,
                SUM(CASE WHEN estado = 'Pendiente' THEN total        ELSE 0 END) AS total_pendiente
            FROM facturas
            WHERE fecha_emision >= ? AND fecha_emision <= ?
        """, (fecha_inicio, fecha_fin))
        totales_row = cursor.fetchone()
        conn.close()

        t = dict(totales_row) if totales_row else {}
        for k in t:
            if t[k] is None:
                t[k] = 0.0
        ivu_total_devengado = (t.get("ivu_estatal", 0) or 0) + (t.get("ivu_municipal", 0) or 0)

        # ── Construir workbook ──────────────────────────────────────────────
        wb = Workbook()
        ws = wb.active
        ws.title = "IVU por Período"

        # Estilos
        COLOR_VERDE   = "3D5A27"
        COLOR_VERDE_L = "C8D9B8"
        COLOR_GRIS    = "F2F2F2"
        COLOR_TOTAL   = "E8F0E0"

        def hdr_font(bold=True, color="FFFFFF", size=10):
            return Font(bold=bold, color=color, size=size, name="Calibri")

        def cell_font(bold=False, size=10, color="000000"):
            return Font(bold=bold, size=size, name="Calibri", color=color)

        def fill(hex_color):
            return PatternFill("solid", fgColor=hex_color)

        def border():
            s = Side(style="thin", color="CCCCCC")
            return Border(left=s, right=s, top=s, bottom=s)

        def money_fmt(ws, row, col, value):
            cell = ws.cell(row=row, column=col, value=round(float(value or 0), 2))
            cell.number_format = '"$"#,##0.00'
            cell.alignment = Alignment(horizontal="right")
            cell.border = border()
            cell.font = cell_font()
            return cell

        # ── Encabezado del reporte ──────────────────────────────────────────
        ws.merge_cells("A1:I1")
        c = ws["A1"]
        c.value = "Evergreen Love — Reporte IVU por Período"
        c.font = Font(bold=True, size=13, color=COLOR_VERDE, name="Calibri")
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 22

        ws.merge_cells("A2:I2")
        c = ws["A2"]
        c.value = f"Período: {fecha_inicio}  al  {fecha_fin}"
        c.font = Font(size=10, color="555555", name="Calibri")
        c.alignment = Alignment(horizontal="center")

        ws.merge_cells("A3:I3")
        c = ws["A3"]
        c.value = f"Generado: {datetime.datetime.now().strftime('%d/%m/%Y %H:%M')}"
        c.font = Font(size=9, color="888888", name="Calibri", italic=True)
        c.alignment = Alignment(horizontal="center")

        ws.row_dimensions[4].height = 6  # separador

        # ── Cabecera de tabla ──────────────────────────────────────────────
        headers = [
            "# Factura", "Fecha", "Cliente",
            "Base Imponible", "IVU Estatal", "IVU Municipal",
            "IVU Total", "Total Factura", "Estado"
        ]
        col_widths = [14, 13, 30, 16, 14, 16, 12, 14, 12]

        for ci, (h, w) in enumerate(zip(headers, col_widths), start=1):
            cell = ws.cell(row=5, column=ci, value=h)
            cell.font = hdr_font()
            cell.fill = fill(COLOR_VERDE)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border()
            ws.column_dimensions[get_column_letter(ci)].width = w

        ws.row_dimensions[5].height = 18

        # ── Filas de datos ─────────────────────────────────────────────────
        estado_colors = {"Pagada": "D6EAD0", "Anulada": "F0F0F0", "Pendiente": "FFF3CD"}

        for ri, row in enumerate(rows, start=6):
            bg = estado_colors.get(row["estado"], "FFFFFF")
            row_fill = fill(bg)

            def txt(col, val, bold=False, align="left"):
                cell = ws.cell(row=ri, column=col, value=val)
                cell.font = cell_font(bold=bold)
                cell.fill = row_fill
                cell.alignment = Alignment(horizontal=align, vertical="center")
                cell.border = border()

            txt(1, row["numero_factura"], align="center")
            txt(2, row["fecha_emision"], align="center")
            txt(3, row["cliente"])
            for ci, field in enumerate(["subtotal","ivu_estatal","ivu_municipal","ivu_total","total"], start=4):
                cell = money_fmt(ws, ri, ci, row[field])
                cell.fill = row_fill
            txt(9, row["estado"], align="center")

        # ── Fila separadora visual ─────────────────────────────────────────
        sep_row = 6 + len(rows)
        ws.row_dimensions[sep_row].height = 8

        # ── Bloque de totales ──────────────────────────────────────────────
        totales = [
            ("Ventas Sujetas a IVU",   t.get("ventas_sujetas",  0)),
            ("IVU Estatal (10.5%)",    t.get("ivu_estatal",     0)),
            ("IVU Municipal (1.0%)",   t.get("ivu_municipal",   0)),
            ("IVU Total Devengado",    ivu_total_devengado),
            ("Total Facturado",        t.get("total_facturado", 0)),
            ("Total Recaudado",        t.get("total_recaudado", 0)),
            ("Total Pendiente",        t.get("total_pendiente", 0)),
        ]

        tr_start = sep_row + 1
        for ti, (label, value) in enumerate(totales):
            tr = tr_start + ti
            # Label (columnas A:G combinadas)
            ws.merge_cells(f"A{tr}:G{tr}")
            lc = ws[f"A{tr}"]
            lc.value = label
            lc.font = Font(bold=True, size=10, name="Calibri", color=COLOR_VERDE)
            lc.fill = fill(COLOR_TOTAL)
            lc.alignment = Alignment(horizontal="right", vertical="center")
            lc.border = border()
            # Valor en columna H
            vc = money_fmt(ws, tr, 8, value)
            vc.font = Font(bold=True, size=10, name="Calibri")
            vc.fill = fill(COLOR_TOTAL)
            # Columna I vacía con mismo fondo
            ec = ws.cell(row=tr, column=9, value="")
            ec.fill = fill(COLOR_TOTAL)
            ec.border = border()
            ws.row_dimensions[tr].height = 17

        # ── Serializar a bytes ─────────────────────────────────────────────
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        filename = f"ivu_evergreen_{fecha_inicio}_{fecha_fin}.xlsx"
        headers_resp = {
            "Content-Disposition": f'attachment; filename="{filename}"'
        }
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=headers_resp
        )

    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl no está instalado. Ejecuta: pip install openpyxl")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al generar Excel: {str(e)}")


@router.get("/gastos")
def listar_gastos(current_user: dict = Depends(get_current_admin)):
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM gastos ORDER BY fecha DESC, id DESC")
        gastos = [dict(row) for row in cursor.fetchall()]
        conn.close()
        return {"status": "success", "data": gastos}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al listar gastos: {str(e)}")

@router.post("/gastos", status_code=201)
def crear_gasto(gasto: GastoCreate, current_user: dict = Depends(get_current_admin)):
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("""
            INSERT INTO gastos (concepto, categoria, monto, fecha, metodo_pago, notas)
            VALUES (?, ?, ?, ?, ?, ?)
        """, (gasto.concepto, gasto.categoria, gasto.monto, gasto.fecha, gasto.metodo_pago, gasto.notas))
        gasto_id = cursor.lastrowid
        conn.commit()
        conn.close()
        return {
            "status": "success",
            "message": "Gasto registrado con éxito",
            "id": gasto_id,
            "concepto": gasto.concepto,
            "categoria": gasto.categoria,
            "monto": gasto.monto,
            "fecha": gasto.fecha,
            "metodo_pago": gasto.metodo_pago,
            "notas": gasto.notas
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al crear gasto: {str(e)}")

@router.delete("/gastos/{gasto_id}")
def eliminar_gasto(gasto_id: int, current_user: dict = Depends(get_current_admin)):
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT id FROM gastos WHERE id = ?", (gasto_id,))
        if not cursor.fetchone():
            conn.close()
            raise HTTPException(status_code=404, detail="Gasto no encontrado")
        cursor.execute("DELETE FROM gastos WHERE id = ?", (gasto_id,))
        conn.commit()
        conn.close()
        return {"status": "success", "message": "Gasto eliminado con éxito"}
    except HTTPException as he:
        raise he
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al eliminar gasto: {str(e)}")


@router.put("/gastos/{gasto_id}")
def actualizar_gasto(
    gasto_id: int,
    data: GastoUpdate,
    current_user: dict = Depends(get_current_admin)
):
    """Actualiza campos de un gasto existente. Solo actualiza los campos que se envían."""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT id FROM gastos WHERE id = ?", (gasto_id,))
        if not cursor.fetchone():
            conn.close()
            raise HTTPException(status_code=404, detail="Gasto no encontrado")

        campos = []
        valores = []
        if data.concepto    is not None: campos.append("concepto = ?");    valores.append(data.concepto)
        if data.categoria   is not None: campos.append("categoria = ?");   valores.append(data.categoria)
        if data.monto       is not None: campos.append("monto = ?");       valores.append(data.monto)
        if data.fecha       is not None: campos.append("fecha = ?");       valores.append(data.fecha)
        if data.metodo_pago is not None: campos.append("metodo_pago = ?"); valores.append(data.metodo_pago)
        if data.notas       is not None: campos.append("notas = ?");       valores.append(data.notas)
        if data.proveedor   is not None: campos.append("proveedor = ?");   valores.append(data.proveedor)

        if not campos:
            conn.close()
            return {"status": "success", "message": "Sin cambios"}

        valores.append(gasto_id)
        cursor.execute(f"UPDATE gastos SET {', '.join(campos)} WHERE id = ?", valores)
        conn.commit()
        conn.close()
        return {"status": "success", "message": "Gasto actualizado correctamente"}
    except HTTPException as he:
        raise he
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al actualizar gasto: {str(e)}")


@router.post("/gastos/{gasto_id}/recibo")
def subir_recibo_gasto(
    gasto_id: int,
    archivo: UploadFile = File(...),
    current_user: dict = Depends(get_current_admin)
):
    ext = os.path.splitext(archivo.filename or "")[1].lower()
    if ext not in ALLOWED_EXTENSIONS:
        raise HTTPException(status_code=400, detail=f"Tipo de archivo no permitido. Solo: {', '.join(ALLOWED_EXTENSIONS)}")

    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT id, recibo_ruta FROM gastos WHERE id = ?", (gasto_id,))
    row = cursor.fetchone()
    if not row:
        conn.close()
        raise HTTPException(status_code=404, detail="Gasto no encontrado")

    # Borrar recibo anterior si existe
    viejo = row["recibo_ruta"]
    if viejo and os.path.exists(viejo):
        try:
            os.remove(viejo)
        except Exception:
            pass

    os.makedirs(RECIBOS_DIR, exist_ok=True)
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    nombre_archivo = f"recibo_{gasto_id}_{timestamp}{ext}"
    ruta_destino = os.path.join(RECIBOS_DIR, nombre_archivo)

    try:
        with open(ruta_destino, "wb") as f:
            shutil.copyfileobj(archivo.file, f)
    except Exception as e:
        conn.close()
        raise HTTPException(status_code=500, detail=f"Error al guardar archivo: {str(e)}")

    cursor.execute("UPDATE gastos SET recibo_ruta = ? WHERE id = ?", (ruta_destino, gasto_id))
    conn.commit()
    conn.close()
    return {"status": "success", "message": "Recibo subido correctamente", "recibo_ruta": nombre_archivo}


@router.get("/gastos/{gasto_id}/recibo")
def ver_recibo_gasto(
    gasto_id: int,
    current_user: dict = Depends(get_current_admin)
):
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT recibo_ruta FROM gastos WHERE id = ?", (gasto_id,))
    row = cursor.fetchone()
    conn.close()

    if not row or not row["recibo_ruta"]:
        raise HTTPException(status_code=404, detail="Este gasto no tiene recibo asociado")

    ruta = row["recibo_ruta"]
    if not os.path.exists(ruta):
        raise HTTPException(status_code=404, detail="Archivo de recibo no encontrado en el servidor")

    media_type, _ = mimetypes.guess_type(ruta)
    media_type = media_type or "application/octet-stream"
    return FileResponse(ruta, media_type=media_type)


@router.delete("/gastos/{gasto_id}/recibo")
def eliminar_recibo_gasto(
    gasto_id: int,
    current_user: dict = Depends(get_current_admin)
):
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT recibo_ruta FROM gastos WHERE id = ?", (gasto_id,))
    row = cursor.fetchone()
    if not row:
        conn.close()
        raise HTTPException(status_code=404, detail="Gasto no encontrado")

    ruta = row["recibo_ruta"]
    if ruta and os.path.exists(ruta):
        try:
            os.remove(ruta)
        except Exception:
            pass

    cursor.execute("UPDATE gastos SET recibo_ruta = NULL WHERE id = ?", (gasto_id,))
    conn.commit()
    conn.close()
    return {"status": "success", "message": "Recibo eliminado"}

@router.get("/gastos/exportar")
def exportar_gastos_excel(
    desde: str = None,
    hasta: str = None,
    categoria: str = None,
    metodo: str = None,
    recibo: str = None,
    current_user: dict = Depends(get_current_admin)
):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import io
    from datetime import date

    conn = get_db_connection()
    cursor = conn.cursor()

    query = "SELECT * FROM gastos WHERE 1=1"
    params = []
    if desde:
        query += " AND fecha >= ?"
        params.append(desde)
    if hasta:
        query += " AND fecha <= ?"
        params.append(hasta)
    if categoria:
        query += " AND categoria = ?"
        params.append(categoria)
    if metodo:
        query += " AND metodo_pago = ?"
        params.append(metodo)
    query += " ORDER BY fecha DESC, id DESC"
    cursor.execute(query, params)
    rows = [dict(r) for r in cursor.fetchall()]
    conn.close()

    if recibo == "con":
        rows = [r for r in rows if r.get("recibo_ruta")]
    elif recibo == "sin":
        rows = [r for r in rows if not r.get("recibo_ruta")]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Gastos"

    moss = "5F7A45"
    light_moss = "EAF0E4"
    header_fill = PatternFill("solid", fgColor=moss)
    total_fill = PatternFill("solid", fgColor=light_moss)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells("A1:G1")
    ws["A1"] = "Evergreen Love — Registro de Gastos"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor=moss)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    periodo_txt = f"{desde or 'inicio'} al {hasta or 'hoy'}"
    ws.merge_cells("A2:G2")
    ws["A2"] = f"Período: {periodo_txt}   |   Generado: {date.today().strftime('%d/%m/%Y')}"
    ws["A2"].font = Font(italic=True, size=10, color="555555")
    ws["A2"].alignment = Alignment(horizontal="center")

    filtros_desc = []
    if categoria:
        filtros_desc.append(f"Categoría: {categoria}")
    if metodo:
        filtros_desc.append(f"Método: {metodo}")
    if recibo == "con":
        filtros_desc.append("Con recibo")
    elif recibo == "sin":
        filtros_desc.append("Sin recibo")
    ws.merge_cells("A3:G3")
    ws["A3"] = ("Filtros: " + " | ".join(filtros_desc)) if filtros_desc else "Sin filtros adicionales"
    ws["A3"].font = Font(italic=True, size=9, color="888888")
    ws["A3"].alignment = Alignment(horizontal="center")

    headers = ["Fecha", "Proveedor", "Concepto", "Categoría", "Método de pago", "Monto", "Recibo"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    ws.row_dimensions[5].height = 22

    total_monto = 0
    con_recibo = 0
    for i, g in enumerate(rows, 6):
        monto = float(g.get("monto") or 0)
        total_monto += monto
        tiene_recibo = bool(g.get("recibo_ruta"))
        if tiene_recibo:
            con_recibo += 1
        row_fill = PatternFill("solid", fgColor="FAFAFA") if i % 2 == 0 else None
        values = [
            g.get("fecha") or "",
            g.get("proveedor") or "",
            g.get("concepto") or "",
            g.get("categoria") or "",
            g.get("metodo_pago") or "",
            monto,
            "Sí" if tiene_recibo else "No",
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.border = border
            cell.alignment = Alignment(vertical="center")
            if row_fill:
                cell.fill = row_fill
            if col == 6:
                cell.number_format = '"$"#,##0.00'
                cell.alignment = Alignment(horizontal="right", vertical="center")
            if col == 7:
                cell.font = Font(color="5F7A45" if tiene_recibo else "CC3333", bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")

    sin_recibo = len(rows) - con_recibo
    total_row = len(rows) + 7
    totales = [
        ("Total gastos:", total_row),
        ("Cantidad de gastos:", total_row + 1),
        ("Con recibo:", total_row + 2),
        ("Sin recibo:", total_row + 3),
    ]
    valores_totales = [total_monto, len(rows), con_recibo, sin_recibo]

    ws.cell(row=total_row - 1, column=1, value="").border = border
    for (label, row), val in zip(totales, valores_totales):
        lc = ws.cell(row=row, column=5, value=label)
        lc.font = Font(bold=True, size=11)
        lc.fill = total_fill
        lc.alignment = Alignment(horizontal="right", vertical="center")
        lc.border = border
        vc = ws.cell(row=row, column=6, value=val)
        vc.font = Font(bold=True, size=11, color=moss)
        vc.fill = total_fill
        vc.border = border
        if row == total_row:
            vc.number_format = '"$"#,##0.00'
            vc.alignment = Alignment(horizontal="right", vertical="center")
        else:
            vc.alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=row, column=7, value="").fill = total_fill

    col_widths = [12, 20, 30, 18, 16, 12, 9]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    desde_str = (desde or "inicio").replace("-", "")
    hasta_str = (hasta or "hoy").replace("-", "")
    filename = f"gastos_evergreen_{desde_str}_{hasta_str}.xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


# ── Enviar factura por email ──────────────────────────────────────────────────

class EnviarEmailSchema(BaseModel):
    email_destino: Optional[str] = None


@router.post("/facturas/{factura_id}/enviar-email")
def enviar_email_factura(
    factura_id: int,
    body: EnviarEmailSchema,
    current_user: dict = Depends(get_current_admin)
):
    """Envía la factura por email al cliente. Requiere SMTP configurado en .env."""
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        # 1. Obtener factura con datos del cliente
        cursor.execute("""
            SELECT f.*,
                   COALESCE(c.nombre, f.cliente_nombre_manual, 'Cliente Directo') AS cliente_nombre,
                   COALESCE(c.email, '') AS cliente_email
            FROM facturas f
            LEFT JOIN clientes c ON f.cliente_id = c.id
            WHERE f.id = ?
        """, (factura_id,))
        row = cursor.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Factura no encontrada")
        factura = dict(row)

        # 2. Resolver email destino: cuerpo del request > BD > error
        email_to = (body.email_destino or "").strip()
        if not email_to:
            email_to = (factura.get("cliente_email") or "").strip()
        if not email_to:
            raise HTTPException(
                status_code=422,
                detail="No hay email disponible. Ingresa uno manualmente o regístralo en el cliente."
            )
        if "@" not in email_to:
            raise HTTPException(status_code=422, detail=f"Email inválido: '{email_to}'")

        # 3. Obtener items de la factura
        cursor.execute("""
            SELECT nombre_producto, cantidad, precio_unitario
            FROM items_factura
            WHERE factura_id = ?
            ORDER BY id ASC
        """, (factura_id,))
        items = [dict(r) for r in cursor.fetchall()]

        # 4. Construir factura_data
        factura_data = {
            "numero_factura":   factura["numero_factura"],
            "fecha_emision":    factura.get("fecha_emision", ""),
            "fecha_vencimiento": factura.get("fecha_vencimiento", ""),
            "cliente_nombre":   factura["cliente_nombre"],
            "estado":           factura.get("estado", "Pendiente"),
            "subtotal":         float(factura.get("subtotal") or 0),
            "ivu_estatal":      float(factura.get("ivu_estatal") or 0),
            "ivu_municipal":    float(factura.get("ivu_municipal") or 0),
            "total":            float(factura.get("total") or 0),
            "notas":            factura.get("notas", "") or "",
            "items":            items,
        }

        # 5. Enviar (foreground para que el admin vea el resultado inmediato)
        ok, mensaje = send_factura_email(email_to, factura_data)
        if not ok:
            raise HTTPException(
                status_code=500,
                detail=f"Error SMTP al enviar el email: {mensaje}"
            )

        return {"status": "success", "email_enviado_a": email_to, "mensaje": mensaje}

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()
